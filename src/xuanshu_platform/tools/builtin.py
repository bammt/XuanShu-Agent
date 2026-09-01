from pathlib import Path
from contextlib import contextmanager
from contextvars import ContextVar
import hashlib
import json
import os
import re
from typing import Any, ClassVar, Literal
import httpx
from crewai.tools import BaseTool
from crewai.mcp import MCPServerHTTP, MCPServerSSE
from crewai.mcp.tool_resolver import MCPToolResolver
from pydantic import BaseModel, Field
from docx import Document
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pypdf import PdfReader
from ..config import settings

_runtime_idempotency_key: ContextVar[str] = ContextVar('xuanshu_runtime_idempotency_key', default='')
_runtime_tool_invocations: ContextVar[dict[str, int] | None] = ContextVar(
    'xuanshu_runtime_tool_invocations', default=None,
)


@contextmanager
def execution_idempotency_scope(key: str):
    token = _runtime_idempotency_key.set(str(key or ''))
    invocation_token = _runtime_tool_invocations.set({})
    try:
        yield
    finally:
        _runtime_tool_invocations.reset(invocation_token)
        _runtime_idempotency_key.reset(token)


def current_idempotency_key() -> str:
    return _runtime_idempotency_key.get()


def tool_idempotency_key(operation: str, payload: Any) -> str:
    """Derive a replay-safe key for one tool call inside a runtime node.

    A CrewAI Agent may call several tools while executing one node.  The node
    key therefore cannot be sent directly to the executor: doing so makes the
    second, different command conflict with the first.  The occurrence counter
    also keeps two intentional calls with identical arguments distinct while
    producing the same sequence of keys when the whole node is retried.
    """
    scope_key = current_idempotency_key()
    if not scope_key:
        return ''
    serialized = json.dumps(
        {'operation': str(operation), 'payload': payload},
        ensure_ascii=False,
        sort_keys=True,
        separators=(',', ':'),
        default=str,
    )
    signature = hashlib.sha256(serialized.encode()).hexdigest()
    invocations = _runtime_tool_invocations.get()
    occurrence = 1
    if invocations is not None:
        occurrence = invocations.get(signature, 0) + 1
        invocations[signature] = occurrence
    material = f'{scope_key}\0{signature}\0{occurrence}'
    return f'xuanshu-tool-{hashlib.sha256(material.encode()).hexdigest()}'

_crewai_server_name = MCPToolResolver._extract_server_name

def safe_mcp_server_name(server_url: str) -> str:
    """Work around CrewAI 1.15.x producing OpenAI-invalid IP prefixes."""
    name = _crewai_server_name(server_url)
    name = re.sub(r'[^A-Za-z0-9_-]', '_', name)
    if not name or not re.match(r'^[A-Za-z_]', name):
        name = f'mcp_{name}'
    return name[:48]

# MCP tools are resolved lazily by CrewAI after these configs are returned.
MCPToolResolver._extract_server_name = staticmethod(safe_mcp_server_name)
from ..services import (app_root_dir, app_session_dir, delete_session_file, resolve_app_file,
                        sync_existing_session_file)

class FileInput(BaseModel): filename:str=Field(description='应用工作目录中的文件名')
class SpreadsheetInput(FileInput):
    sheet_name: str | None = Field(default=None, description='可选工作表名称；不填写时读取所有工作表')
    start_row: int = Field(default=1, ge=1, le=1_048_576, description='开始读取的行号，从 1 开始')
    max_rows: int = Field(default=100, ge=1, le=500, description='每个工作表本次最多读取的行数')
    max_columns: int = Field(default=30, ge=1, le=200, description='每个工作表本次最多读取的列数')
class CodeInput(BaseModel): code:str=Field(description='需要执行的 Python 代码')
class CommandInput(BaseModel): command:str=Field(description='需要在应用隔离工作目录执行的 shell 命令')
class SkillCommandInput(BaseModel):
    skill_id: str = Field(description='当前智能体已绑定的 Skill ID')
    command: str = Field(description='从该 Skill 包目录执行的 shell 命令')
class DynamicToolInput(BaseModel):
    arguments: dict[str, Any] = Field(default_factory=dict, description='按工具说明提供的参数对象')


class ExecutionReceiptStore:
    def __init__(self) -> None:
        self.items: list[dict] = []


class AskUserRequestStore:
    """Per-run bridge from an Agent tool call to the durable runtime.

    The tool itself never waits on a worker thread.  It records the request and
    the runtime turns the current run into ``waiting_input`` immediately after
    the CrewAI call returns.  This keeps one worker available for other users
    while the conversation is paused.
    """

    def __init__(self) -> None:
        self.request: dict[str, Any] | None = None
        self.error: str | None = None
        self.failure: dict[str, Any] | None = None

    def set(self, request: dict[str, Any]) -> None:
        if self.request is None:
            self.request = request

    def consume(self) -> dict[str, Any] | None:
        request = self.request
        self.request = None
        return request

    def set_error(self, message: str, *, tool_name: str = 'ask_user',
                  arguments: dict[str, Any] | None = None) -> None:
        if self.error is None:
            self.error = str(message)
            self.failure = {
                'tool_name': str(tool_name or 'ask_user'),
                'arguments': dict(arguments or {}),
                'error': str(message),
            }

    def consume_failure(self) -> dict[str, Any] | None:
        failure = self.failure
        self.failure = None
        self.error = None
        return failure

    def consume_error(self) -> str | None:
        failure = self.consume_failure()
        return str(failure.get('error') or '') if failure else None


class AskUserInput(BaseModel):
    question: str = Field(description='需要向最终用户提出的具体问题')
    input_name: str = Field(
        default='',
        description=('可选；只能填写当前应用已声明运行输入的机器字段名，不能填写 text、file、image 等类型；'
                     '通用业务追问必须留空'),
    )
    input_type: Literal['text', 'long_text', 'file', 'image', 'number', 'boolean', 'json'] | None = Field(
        default=None, description='可省略；这里才填写 text、file、image 等字段类型，且必须与 input_name 的声明一致'
    )
    required: bool = Field(default=True, description='该信息是否必须提供')


class AskUserTool(BaseTool):
    name: str = 'ask_user'
    description: str = (
        '当继续执行需要用户补充信息时调用。问题必须具体、一次只询问当前最必要的信息。'
        '调用后当前运行会暂停，用户回复后从当前节点继续。input_name 是可选的界面和状态提示，'
        '通用追问不要填写。用户可在回复时同时上传已声明的文件，文本和附件由会话通道独立持久化；'
        '不要把控制状态或 JSON 写进最终答复。'
    )
    args_schema: type[BaseModel] = AskUserInput
    result_as_answer: bool = True
    request_store: Any = Field(exclude=True)
    allowed_inputs: dict[str, dict[str, Any]] = Field(default_factory=dict, exclude=True)

    # A few OpenAI-compatible models fill an optional ``input_name`` with the
    # schema type (usually ``text``) even when the question is generic. These
    # transport placeholders are not application variables.
    GENERIC_INPUT_NAMES: ClassVar[frozenset[str]] = frozenset({
        'text', 'message', 'user_message', 'chat_message', 'input', 'user_input',
    })

    def _run(self, question: str, input_name: str = '', input_type: str | None = None,
             required: bool = True) -> str:
        arguments = {
            'question': str(question or ''),
            'input_name': str(input_name or ''),
            'input_type': input_type,
            'required': bool(required),
        }
        question = str(question or '').strip()
        if not question:
            message = '无法向用户提问：question 不能为空。'
            self.request_store.set_error(message, tool_name=self.name, arguments=arguments)
            return message
        input_name = str(input_name or '').strip()
        contract = self.allowed_inputs.get(input_name) if input_name else None
        if input_name and contract is None and input_name.casefold() in self.GENERIC_INPUT_NAMES:
            input_name = ''
            contract = None
        if input_name and contract is None:
            allowed = '、'.join(self.allowed_inputs) or '无'
            message = (f'无法向用户提问：input_name={input_name} 不在当前应用的运行输入契约中。'
                       f'当前可用字段：{allowed}。')
            self.request_store.set_error(message, tool_name=self.name, arguments=arguments)
            return message
        requested_type = str(input_type or '').strip()
        normalized_type = str((contract or {}).get('input_type') or requested_type or 'text').strip()
        if contract and requested_type and requested_type != contract.get('input_type'):
            # ``text`` is the generic chat value used by some compatible
            # models; it is valid for a long_text field because the platform
            # still stores the complete user message under that field.
            if requested_type == 'text' and contract.get('input_type') == 'long_text':
                requested_type = str(contract.get('input_type'))
            else:
                message = (f'无法向用户提问：字段 {input_name} 的类型必须是 '
                           f"{contract.get('input_type')}，不能是 {requested_type}。")
                self.request_store.set_error(message, tool_name=self.name, arguments=arguments)
                return message
        file_inputs = [
            name for name, item in self.allowed_inputs.items()
            if str(item.get('input_type') or '') in {'file', 'image'}
        ]
        self.request_store.set({
            'question': question,
            'input_name': input_name,
            'input_type': normalized_type,
            'required': bool((contract or {}).get('required', required)),
            'accepts_files': bool(file_inputs),
            'file_input_names': file_inputs,
        })
        return '已记录用户问题。请停止当前节点，等待用户回复后再继续。'
class DocumentReaderTool(BaseTool):
    name:str='read_document'; description:str='读取应用工作目录中的 DOCX 或 PDF 文件正文'; args_schema:type[BaseModel]=FileInput
    workspace_id:int; app_id:int; execution_scope:str; app_kind:str='crew'
    def _run(self,filename:str)->str:
        root=app_session_dir(self.workspace_id,self.app_id,self.execution_scope,self.app_kind).resolve()
        try: path=resolve_app_file(root,filename)
        except ValueError: return '文件不存在'
        if not path.is_file(): return '文件不存在'
        if path.suffix.lower()=='.docx': return '\n'.join(p.text for p in Document(path).paragraphs)
        if path.suffix.lower()=='.pdf': return '\n'.join((p.extract_text() or '') for p in PdfReader(path).pages)
        return '仅支持 DOCX 和 PDF'
class SpreadsheetReaderTool(BaseTool):
    name: str = 'read_spreadsheet'
    description: str = (
        '读取应用工作目录中的 XLSX 工作簿。可读取全部或指定工作表，保留单元格值和公式；'
        '大型工作表可通过 start_row 分段读取。'
    )
    args_schema: type[BaseModel] = SpreadsheetInput
    workspace_id: int
    app_id: int
    execution_scope: str
    app_kind: str = 'crew'

    @staticmethod
    def _cell_text(value: Any) -> str:
        if value is None:
            return ''
        if hasattr(value, 'isoformat'):
            try:
                value = value.isoformat()
            except (TypeError, ValueError):
                pass
        return str(value).replace('|', '\\|').replace('\r\n', '<br>').replace('\n', '<br>').replace('\r', '<br>')

    def _run(self, filename: str, sheet_name: str | None = None, start_row: int = 1,
             max_rows: int = 100, max_columns: int = 30) -> str:
        root = app_session_dir(self.workspace_id, self.app_id, self.execution_scope, self.app_kind).resolve()
        try:
            path = resolve_app_file(root, filename)
        except ValueError:
            return '文件不存在'
        if not path.is_file():
            return '文件不存在'
        if path.suffix.lower() != '.xlsx':
            return '仅支持 XLSX 文件'
        try:
            workbook = load_workbook(path, read_only=True, data_only=False)
        except Exception as exc:
            return f'无法读取 XLSX 文件：{exc}'
        try:
            if sheet_name and sheet_name not in workbook.sheetnames:
                return f'工作表不存在。当前工作表：{"、".join(workbook.sheetnames)}'
            worksheets = [workbook[sheet_name]] if sheet_name else list(workbook.worksheets)
            lines = [f'工作簿：{path.name}', f'工作表：{"、".join(workbook.sheetnames)}']
            output_limit = 20_000
            output_size = sum(len(item) + 1 for item in lines)
            output_truncated = False

            def append(line: str) -> bool:
                nonlocal output_size, output_truncated
                line_size = len(line) + 1
                if output_size + line_size > output_limit:
                    output_truncated = True
                    return False
                lines.append(line)
                output_size += line_size
                return True

            for worksheet in worksheets:
                column_count = min(max(worksheet.max_column, 1), max_columns)
                last_row = min(worksheet.max_row, start_row + max_rows - 1)
                if not append(f'\n## 工作表：{worksheet.title}'):
                    break
                append(
                    f'范围：{worksheet.calculate_dimension()}；本次读取第 {start_row}-{last_row} 行，'
                    f'最多 {column_count} 列'
                )
                headers = ['行号', *(get_column_letter(index) for index in range(1, column_count + 1))]
                if not append('| ' + ' | '.join(headers) + ' |'):
                    break
                if not append('| ' + ' | '.join('---' for _ in headers) + ' |'):
                    break
                has_data = False
                if start_row <= last_row:
                    for row_number, row in enumerate(
                        worksheet.iter_rows(
                            min_row=start_row, max_row=last_row, min_col=1, max_col=column_count,
                            values_only=True,
                        ),
                        start=start_row,
                    ):
                        values = [self._cell_text(value) for value in row]
                        if not any(values):
                            continue
                        has_data = True
                        if not append('| ' + ' | '.join([str(row_number), *values]) + ' |'):
                            break
                if output_truncated:
                    break
                if not has_data:
                    append('（指定范围内无数据）')
                if worksheet.max_row > last_row:
                    append(
                        f'（工作表还有更多行；使用 sheet_name="{worksheet.title}"、'
                        f'start_row={last_row + 1} 继续读取。）'
                    )
                if worksheet.max_column > column_count:
                    append(f'（工作表共有 {worksheet.max_column} 列，本次只读取前 {column_count} 列。）')
            if output_truncated:
                lines.append('（输出达到长度限制；请指定 sheet_name 和 start_row 分段读取。）')
            return '\n'.join(lines)
        finally:
            workbook.close()
class SandboxedPythonTool(BaseTool):
    name:str='execute_python'; description:str=(
        '在当前应用隔离工作目录执行 Python。已绑定 Skill 通过 '
        'xuanshu_skills.<命名空间> 独立导入，$XUANSHU_SKILL_MAP 提供命名空间到包根目录的映射；'
        '也允许按任务需要编写新代码、联网下载字体和依赖。'
    ); args_schema:type[BaseModel]=CodeInput
    workspace_id:int; app_id:int
    execution_scope:str
    app_kind:str='crew'
    environment: dict[str, str] = Field(default_factory=dict, exclude=True)
    skill_roots: list[str] = Field(default_factory=list, exclude=True)
    receipt_store: Any = Field(default=None, exclude=True)
    receipt_store: Any = Field(default=None, exclude=True)

    def _execute_payload(self, payload: dict) -> dict:
        executor_payload = {
            **payload,
            'workspace_id': self.workspace_id,
            'application_id': self.app_id,
            'application_kind': self.app_kind,
            'execution_scope': self.execution_scope,
            'skill_roots': self.skill_roots,
            'environment': self.environment,
        }
        response = httpx.post(
            f'{settings.executor_url}/execute',
            json={**executor_payload,
                  'idempotency_key': tool_idempotency_key('executor.execute', executor_payload)},
            headers={'X-Executor-Token': settings.executor_shared_secret},
            timeout=settings.code_timeout_seconds + 10,
        )
        response.raise_for_status()
        data = response.json()
        for item in data.get('files', []):
            sync_existing_session_file(
                self.workspace_id, self.app_id, self.execution_scope, item['path'], self.app_kind,
            )
        for path in data.get('deleted_files', []):
            delete_session_file(
                self.workspace_id, self.app_id, self.execution_scope, path, self.app_kind,
            )
        if self.receipt_store is not None:
            program = str(payload.get('code') or payload.get('command') or '')
            self.receipt_store.items.append({
                'tool': self.name,
                'program_sha256': hashlib.sha256(program.encode()).hexdigest(),
                'exit_code': data.get('exit_code'),
                'files': [item['path'] for item in data.get('files', [])],
                'deleted_files': list(data.get('deleted_files', [])),
            })
        return data

    def execute(self, code: str) -> dict:
        return self._execute_payload({'code': code})

    def _run(self,code:str)->str:
        try:
            data = self.execute(code)
            delivered = [item['path'] for item in data.get('files', [])]
            delivery = f"\n生成或修改文件：{', '.join(delivered)}" if delivered else ''
            return f"exit_code={data['exit_code']}\nstdout:\n{data['stdout']}\nstderr:\n{data['stderr']}{delivery}"
        except Exception as exc: return f'代码执行失败：{exc}'

class SandboxedCommandTool(SandboxedPythonTool):
    name:str='execute_command'
    description:str=(
        '在当前应用隔离工作目录执行 shell 命令。可通过 $XUANSHU_SKILL_MAP 精确找到每个完整 '
        'Skill 包，允许联网下载字体、安装依赖和运行命令；只能读写该应用目录。'
    )
    args_schema:type[BaseModel]=CommandInput

    def _run(self,command:str)->str:
        try:
            data = self._execute_payload({'command': command})
            delivered = [item['path'] for item in data.get('files', [])]
            delivery = f"\n生成或修改文件：{', '.join(delivered)}" if delivered else ''
            return f"exit_code={data['exit_code']}\nstdout:\n{data['stdout']}\nstderr:\n{data['stderr']}{delivery}"
        except Exception as exc: return f'命令执行失败：{exc}'


class SkillCommandTool(SandboxedPythonTool):
    name:str='execute_skill_script'
    description:str=(
        '需要以 Skill 包为当前目录时使用。传入已绑定 Skill ID 和命令；'
        '这是便捷执行方式，不限制使用 execute_python 或 execute_command 完成其他代码工作。'
    )
    args_schema:type[BaseModel]=SkillCommandInput
    skill_entries:dict[str,dict]=Field(default_factory=dict,exclude=True)
    def _run(self,skill_id:str,command:str)->str:
        requested = str(skill_id or '').strip()
        entry = self.skill_entries.get(requested)
        if not entry:
            # Models frequently use the human-readable Skill slug/name even
            # though the platform schema exposes the numeric resource ID.
            # Resolve only a unique alias; never guess between two Skills.
            aliases = []
            for bound_id, candidate in self.skill_entries.items():
                values = {
                    str(bound_id),
                    str(candidate.get('skill_id') or ''),
                    str(candidate.get('slug') or ''),
                    str(candidate.get('name') or ''),
                    str(candidate.get('python_namespace') or ''),
                }
                if requested and requested.casefold() in {value.casefold() for value in values if value}:
                    aliases.append((bound_id, candidate))
            if len(aliases) == 1:
                skill_id, entry = aliases[0]
        if not entry:
            return f'Skill 脚本执行失败：Skill {requested} 未绑定到当前智能体'
        relative=Path(str(entry.get('discovery_root') or ''))/str(entry.get('slug') or '')
        root=app_root_dir(self.workspace_id,self.app_id,self.app_kind).resolve()
        skill_root=(root/relative).resolve()
        if root not in skill_root.parents or not (skill_root/'SKILL.md').is_file():
            return f'Skill 脚本执行失败：Skill {skill_id} 快照不存在'
        script_files=[path.relative_to(skill_root).as_posix() for path in (skill_root/'scripts').rglob('*')
                      if path.is_file() and not path.is_symlink()]
        matched_scripts=[path for path in script_files if path in command or Path(path).with_suffix('').as_posix().replace('/','.') in command]
        try:
            receipt_index = len(self.receipt_store.items) if self.receipt_store is not None else 0
            data=self._execute_payload({
                'command':command,'working_directory':relative.as_posix(),'publish_outputs':True,
            })
            receipt={
                'skill_id':str(skill_id),'slug':entry.get('slug'),'digest':entry.get('digest'),
                'command_sha256':hashlib.sha256(command.encode()).hexdigest(),
                'matched_scripts':matched_scripts,
                'exit_code':data.get('exit_code'),'files':[item['path'] for item in data.get('files',[])],
            }
            if self.receipt_store is not None:
                if len(self.receipt_store.items) > receipt_index:
                    self.receipt_store.items[receipt_index].update(receipt)
                else:
                    self.receipt_store.items.append(receipt)
            receipt_root=root/'runtime'/'executions'; receipt_root.mkdir(parents=True,exist_ok=True)
            receipt_path=receipt_root/f'{len(list(receipt_root.glob("*.json")))+1:06d}-{receipt["command_sha256"][:12]}.json'
            receipt_path.write_text(json.dumps(receipt,ensure_ascii=False,indent=2),encoding='utf-8')
            delivered=receipt['files']
            delivery=f"\n生成或修改文件：{', '.join(delivered)}" if delivered else ''
            return f"exit_code={data['exit_code']}\nstdout:\n{data['stdout']}\nstderr:\n{data['stderr']}{delivery}"
        except Exception as exc:
            return f'Skill 脚本执行失败：{exc}'

class ConfiguredHttpTool(BaseTool):
    name: str
    description: str
    args_schema: type[BaseModel] = DynamicToolInput
    endpoint: str
    method: str = 'POST'
    headers: dict[str, str] = Field(default_factory=dict)
    request_template: dict = Field(default_factory=dict)
    response_path: str = ''

    def _run(self, arguments: dict[str, Any] | None = None) -> str:
        values = arguments or {}
        try:
            endpoint = self.endpoint.format(**values)
            payload = {**self.request_template, **values}
            method = self.method.upper()
            headers = dict(self.headers)
            if current_idempotency_key() and 'Idempotency-Key' not in headers:
                headers['Idempotency-Key'] = tool_idempotency_key('http.request', {
                    'method': method,
                    'endpoint': endpoint,
                    'payload': payload,
                })
            kwargs = {'headers': headers, 'timeout': 30.0}
            if method == 'GET': kwargs['params'] = payload
            else: kwargs['json'] = payload
            response = httpx.request(method, endpoint, **kwargs)
            response.raise_for_status()
            data: Any = response.json() if 'json' in response.headers.get('content-type', '') else response.text
            for part in filter(None, self.response_path.split('.')):
                data = data[int(part)] if isinstance(data, list) else data[part]
            return data if isinstance(data, str) else json.dumps(data, ensure_ascii=False)
        except Exception as exc:
            return f'HTTP 工具调用失败：{exc}'

class ConfiguredPythonTool(BaseTool):
    name: str
    description: str
    args_schema: type[BaseModel] = DynamicToolInput
    workspace_id: int
    app_id: int
    execution_scope: str
    app_kind: str = 'crew'
    source_code: str
    environment: dict[str, str] = Field(default_factory=dict, exclude=True)
    skill_roots: list[str] = Field(default_factory=list, exclude=True)

    def _run(self, arguments: dict[str, Any] | None = None) -> str:
        invocation = (self.source_code + '\n\n'
                      + f'__xuanshu_arguments = {json.dumps(arguments or {}, ensure_ascii=False)!r}\n'
                      + 'import json as __xuanshu_json\n'
                      + '__xuanshu_arguments = __xuanshu_json.loads(__xuanshu_arguments)\n'
                      + 'if "run" in globals():\n    print(run(**__xuanshu_arguments))\n'
                      + 'elif "main" in globals():\n    print(main(**__xuanshu_arguments))\n'
                      + 'else:\n    raise RuntimeError("Python Tool 必须定义 run(**kwargs) 或 main(**kwargs)")\n')
        return SandboxedPythonTool(workspace_id=self.workspace_id, app_id=self.app_id,
                                   execution_scope=self.execution_scope, app_kind=self.app_kind,
                                   environment=self.environment, skill_roots=self.skill_roots,
                                   receipt_store=self.receipt_store).run(code=invocation)

def configured_capabilities(
    workspace_id: int,
    app_id: int,
    plugins: list[dict],
    *,
    app_kind: str = 'crew',
    execution_scope: str,
    skill_roots: list[str] | None = None,
    receipt_store: ExecutionReceiptStore | None = None,
) -> tuple[list[BaseTool], list, list[str]]:
    tools: list[BaseTool] = []
    mcps: list = []
    apps: list[str] = []
    for item in plugins:
        if not item.get('enabled', True):
            continue
        raw_name = str(item.get('tool_name') or f"workspace_tool_{item['id']}")
        function_name = re.sub(r'[^A-Za-z0-9_-]', '_', raw_name)
        if not function_name or not re.match(r'^[A-Za-z_]', function_name):
            function_name = f'tool_{function_name}'
        common = {'name': function_name[:64],
                  'description': item.get('description') or item.get('name', '工作空间工具')}
        if item.get('kind') == 'http' and item.get('endpoint'):
            tools.append(ConfiguredHttpTool(**common, endpoint=item['endpoint'], method=item.get('method', 'POST'),
                                            headers=item.get('headers') or {}, request_template=item.get('request_template') or {},
                                            response_path=item.get('response_path', '')))
        elif item.get('kind') == 'python' and item.get('source_code'):
            names = [str(name) for name in item.get('env_vars', [])]
            environment = {name: os.environ[name] for name in names if name in os.environ}
            tools.append(ConfiguredPythonTool(**common, workspace_id=workspace_id, app_id=app_id, app_kind=app_kind,
                                              execution_scope=execution_scope,
                                              source_code=item['source_code'], environment=environment,
                                              skill_roots=skill_roots or [], receipt_store=receipt_store))
        elif item.get('kind') in {'mcp_http', 'mcp_sse'} and item.get('server_url'):
            headers = dict(item.get('headers') or {})
            if item.get('auth_token'):
                headers[item.get('auth_header') or 'Authorization'] = item['auth_token']
            common_mcp = {
                'url': item['server_url'],
                'headers': headers or None,
                'cache_tools_list': bool(item.get('cache_tools_list')),
            }
            mcps.append(MCPServerHTTP(**common_mcp, streamable=True) if item['kind'] == 'mcp_http'
                        else MCPServerSSE(**common_mcp))
        elif item.get('kind') == 'app' and item.get('app_slug'):
            if not settings.crewai_platform_integration_token:
                raise RuntimeError('Connected App 未配置 CREWAI_PLATFORM_INTEGRATION_TOKEN')
            apps.append(item['app_slug'])
    return tools, mcps, apps

def configured_tools(workspace_id: int, app_id: int, plugins: list[dict], *, app_kind: str = 'crew',
                     execution_scope: str = 'legacy') -> list[BaseTool]:
    """Compatibility helper for code paths that only consume ordinary tools."""
    return configured_capabilities(
        workspace_id, app_id, plugins, app_kind=app_kind, execution_scope=execution_scope,
    )[0]
def builtin_tools(workspace_id: int, app_id: int, *, app_kind: str = 'crew', include_code: bool = True,
                  execution_scope: str,
                  skill_entries: dict[str, dict] | None = None, skill_roots: list[str] | None = None,
                  receipt_store: ExecutionReceiptStore | None = None,
                  ask_user_store: AskUserRequestStore | None = None,
                  ask_user_inputs: dict[str, dict[str, Any]] | None = None) -> list[BaseTool]:
    reader_options = {
        'workspace_id': workspace_id,
        'app_id': app_id,
        'execution_scope': execution_scope,
        'app_kind': app_kind,
    }
    tools: list[BaseTool] = [DocumentReaderTool(**reader_options), SpreadsheetReaderTool(**reader_options)]
    if ask_user_store is not None:
        tools.append(AskUserTool(request_store=ask_user_store,
                                 allowed_inputs=ask_user_inputs or {}))
    if include_code:
        tools.extend([
            SandboxedPythonTool(workspace_id=workspace_id, app_id=app_id, execution_scope=execution_scope,
                                app_kind=app_kind,
                                skill_roots=skill_roots or [], receipt_store=receipt_store),
            SandboxedCommandTool(workspace_id=workspace_id, app_id=app_id, execution_scope=execution_scope,
                                 app_kind=app_kind,
                                 skill_roots=skill_roots or [], receipt_store=receipt_store),
            SkillCommandTool(workspace_id=workspace_id, app_id=app_id, execution_scope=execution_scope,
                             app_kind=app_kind,
                             skill_entries=skill_entries or {}, skill_roots=skill_roots or [],
                             receipt_store=receipt_store),
        ])
    return tools
